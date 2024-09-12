import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import os

# Obter o diretório do script atual
script_dir = os.path.dirname(os.path.abspath(__file__))

# Construir o caminho completo para o arquivo CSV
csv_path = os.path.join(script_dir, 'fluxo_de_caixa.csv')

# Verificar se o arquivo existe
if not os.path.exists(csv_path):
    print(f"O arquivo {csv_path} não foi encontrado.")
    print("Por favor, crie o arquivo CSV antes de executar este script.")
    exit()

# Carregar dados
df = pd.read_csv(csv_path)

print("Criando gráfico...")
plt.figure(figsize=(6, 3))
plt.plot(df['Data'], df['Saldo'], marker='o')
plt.title('Evolução do Saldo')
plt.xlabel('Data')
plt.ylabel('Saldo')
plt.xticks(rotation=45)
plt.tight_layout()

print("Salvando gráfico...")
img_buffer = BytesIO()
plt.savefig(img_buffer, format='png', dpi=300, bbox_inches='tight')
img_buffer.seek(0)

print("Criando workbook Excel...")
wb = Workbook()
ws = wb.active
ws.title = "Dashboard Fluxo de Caixa"

print("Adicionando título...")
ws.merge_cells('A1:J1')
ws['A1'] = "Dashboard de Fluxo de Caixa"
ws['A1'].font = Font(size=20, bold=True)
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

print("Adicionando dados...")
headers = ["Data", "Descrição", "Receitas", "Despesas", "Saldo"]
for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=3, column=col, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')

for row, data in enumerate(df.itertuples(index=False), start=4):
    for col, value in enumerate(data, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        if col in [1, 2]:  # Data e Descrição
            cell.alignment = Alignment(horizontal='left', vertical='center')
        else:  # Colunas numéricas
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal='center', vertical='center')

print("Ajustando largura das colunas...")
ws.column_dimensions['A'].width = 12  # Data
ws.column_dimensions['B'].width = 25  # Descrição
for col in range(3, 6):  # Colunas numéricas
    ws.column_dimensions[get_column_letter(col)].width = 15

print("Adicionando bordas...")
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

for row in ws['A3:E' + str(ws.max_row)]:
    for cell in row:
        cell.border = thin_border

print("Adicionando gráfico...")
img = Image(img_buffer)
ws.add_image(img, 'G3')

print("Ajustando altura das linhas...")
for row in range(1, ws.max_row + 1):
    ws.row_dimensions[row].height = 20

print("Salvando arquivo Excel...")
excel_path = os.path.join(script_dir, "dashboard_fluxo_de_caixa_novo.xlsx")
wb.save(excel_path)
print(f"Arquivo Excel salvo em: {excel_path}")

# Verificação adicional
if os.path.exists(excel_path):
    print(f"O arquivo foi criado com sucesso. Tamanho: {os.path.getsize(excel_path)} bytes")
else:
    print("Erro: O arquivo não foi criado!")
